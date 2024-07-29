import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

# Get yesterday's date and format it
yesterday = datetime.now() - timedelta(1)
yesterday_str = yesterday.strftime('%Y-%m-%d')

# Read the CSV file
file_name = f"{yesterday_str}_articles.csv"
data = pd.read_csv(file_name)

print(data.columns)

wb = Workbook()
wb.remove(wb.active)  

def add_bulletin_to_sheet(ws, bulletin, index):
    ws.merge_cells('A1:B1')
    ws['A1'] = f'Bulletin de sécurité  {index + 1} '
    ws['A1'].font = Font(bold=True)
   
    mapping = {
        'Titre': 'title',
        'Date de publication': 'published_date',
        'CVE': 'cve_identifiers',
        'CVSS': 'cvss_scores',
        'Descriptions': 'resume',
        'Impact': 'impacts',
        'Produits impactées': 'impacted_products',
        'Recommandations': 'recommendations',
        'Référence': 'source'
    }
    
    row = 2
    for key, value in mapping.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = bulletin[value]
        row += 1

# Iterate over each row in the DataFrame and create a sheet for each bulletin
for index, row in data.iterrows():
    # Create a new sheet for each bulletin, named by its index
    ws = wb.create_sheet(title=f'Bulletin {index+1}')
    add_bulletin_to_sheet(ws, row, index)

# Save the workbook with a filename based on yesterday's date
excel_file_name = f"{yesterday_str}_security_bulletins.xlsx"
wb.save(excel_file_name)
